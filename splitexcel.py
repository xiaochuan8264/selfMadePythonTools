import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

file = input(r'文件绝对路径：')
content = pd.read_excel(file,header=0)

writer_file = 'splitedExcel.xlsx'
writer = pd.ExcelWriter(writer_file)

splitbase = input('依据哪个字段进行拆分：') #使用pandas的groupby分组数据，并导出到不同表格
res = content.groupby(splitbase)

for i in res:
    i[1].to_excel(writer, index=False, sheet_name=i[0])
writer.save()

workbook = openpyxl.load_workbook('splitedExcel.xlsx')
sheets = workbook.sheetnames

def proceessOneSheet(sheet):
    rows = sheet.max_row
    cols = sheet.max_column
    aleft = Alignment(horizontal='left', vertical='center') # 水平居左，垂直居中
    left, right, top, bottom = [Side(style='thin',color='000000')]*4 #上下左右边框黑色，细框
    for row in range(1,rows+1):
        for col in range(1,cols+1):
            sheet.cell(row=row, column=col).alignment = aleft
            sheet.cell(row=row, column=col).border = border

for sheet_name in sheets:
    sheet = workbook[sheet_name]
    print("处理表单：%s" % sheet_name)
    proceessOneSheet(sheet)
workbook.save('splitedExcel.xlsx')